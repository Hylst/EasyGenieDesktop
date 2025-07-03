"""Export System for Data Export and Format Conversion.

This module provides comprehensive export capabilities for various data formats
including documents, spreadsheets, presentations, and custom formats.
"""

import json
import csv
import xml.etree.ElementTree as ET
from typing import Dict, List, Optional, Any, Union, Callable, Tuple
from dataclasses import dataclass, field
from enum import Enum
from datetime import datetime
from pathlib import Path
import io
import base64
import zipfile
import tempfile

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
except ImportError:
    canvas = None
    SimpleDocTemplate = None

try:
    from docx import Document
    from docx.shared import Inches
except ImportError:
    Document = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    Workbook = None

try:
    from pptx import Presentation
    from pptx.util import Inches as PptxInches
except ImportError:
    Presentation = None


class ExportFormat(Enum):
    """Supported export formats."""
    # Text formats
    TXT = "txt"
    CSV = "csv"
    JSON = "json"
    XML = "xml"
    YAML = "yaml"
    MARKDOWN = "md"
    
    # Document formats
    PDF = "pdf"
    DOCX = "docx"
    RTF = "rtf"
    HTML = "html"
    
    # Spreadsheet formats
    XLSX = "xlsx"
    XLS = "xls"
    ODS = "ods"
    
    # Presentation formats
    PPTX = "pptx"
    PPT = "ppt"
    ODP = "odp"
    
    # Archive formats
    ZIP = "zip"
    TAR = "tar"
    
    # Custom formats
    CUSTOM = "custom"


class ExportQuality(Enum):
    """Export quality settings."""
    DRAFT = "draft"
    STANDARD = "standard"
    HIGH = "high"
    PRINT = "print"


class ExportStatus(Enum):
    """Export operation status."""
    PENDING = "pending"
    PROCESSING = "processing"
    COMPLETED = "completed"
    FAILED = "failed"
    CANCELLED = "cancelled"


class DataType(Enum):
    """Data types for export."""
    TEXT = "text"
    TABLE = "table"
    LIST = "list"
    TREE = "tree"
    CHART = "chart"
    IMAGE = "image"
    MIXED = "mixed"


@dataclass
class ExportConfig:
    """Export configuration."""
    # Basic settings
    format: ExportFormat = ExportFormat.PDF
    quality: ExportQuality = ExportQuality.STANDARD
    
    # File settings
    filename: str = ""
    output_directory: Path = Path(".")
    overwrite_existing: bool = False
    
    # Content settings
    include_metadata: bool = True
    include_timestamps: bool = True
    include_headers: bool = True
    include_footers: bool = False
    
    # Formatting settings
    page_size: str = "A4"  # A4, Letter, Legal, etc.
    orientation: str = "portrait"  # portrait, landscape
    margins: Dict[str, float] = field(default_factory=lambda: {
        "top": 1.0, "bottom": 1.0, "left": 1.0, "right": 1.0
    })
    
    # Font settings
    font_family: str = "Arial"
    font_size: int = 12
    line_spacing: float = 1.2
    
    # Color settings
    use_colors: bool = True
    color_scheme: str = "default"
    
    # Compression settings
    compress: bool = False
    compression_level: int = 6  # 0-9
    
    # Security settings
    password_protect: bool = False
    password: str = ""
    
    # Custom settings
    custom_options: Dict[str, Any] = field(default_factory=dict)


@dataclass
class ExportData:
    """Data to be exported."""
    # Content
    title: str = ""
    content: Any = None
    data_type: DataType = DataType.TEXT
    
    # Metadata
    author: str = ""
    subject: str = ""
    keywords: List[str] = field(default_factory=list)
    created_date: datetime = field(default_factory=datetime.now)
    
    # Structure
    sections: List[Dict[str, Any]] = field(default_factory=list)
    attachments: List[Path] = field(default_factory=list)
    
    # Custom data
    custom_fields: Dict[str, Any] = field(default_factory=dict)


@dataclass
class ExportResult:
    """Export operation result."""
    # Status
    status: ExportStatus = ExportStatus.PENDING
    success: bool = False
    
    # Output
    output_path: Optional[Path] = None
    file_size: int = 0
    
    # Timing
    start_time: datetime = field(default_factory=datetime.now)
    end_time: Optional[datetime] = None
    duration: float = 0.0
    
    # Error handling
    error_message: str = ""
    warnings: List[str] = field(default_factory=list)
    
    # Statistics
    pages_exported: int = 0
    items_exported: int = 0
    
    # Metadata
    export_config: Optional[ExportConfig] = None
    export_data: Optional[ExportData] = None


class BaseExporter:
    """Base class for all exporters."""
    
    def __init__(self, format: ExportFormat):
        """Initialize exporter.
        
        Args:
            format: Export format
        """
        self.format = format
        self.supported_data_types = [DataType.TEXT]
    
    def can_export(self, data_type: DataType) -> bool:
        """Check if exporter can handle data type.
        
        Args:
            data_type: Data type to check
            
        Returns:
            bool: True if supported
        """
        return data_type in self.supported_data_types
    
    def export(self, data: ExportData, config: ExportConfig) -> ExportResult:
        """Export data.
        
        Args:
            data: Data to export
            config: Export configuration
            
        Returns:
            ExportResult: Export result
        """
        raise NotImplementedError("Subclasses must implement export method")
    
    def validate_config(self, config: ExportConfig) -> List[str]:
        """Validate export configuration.
        
        Args:
            config: Configuration to validate
            
        Returns:
            List[str]: List of validation errors
        """
        errors = []
        
        # Check output directory
        if not config.output_directory.exists():
            try:
                config.output_directory.mkdir(parents=True, exist_ok=True)
            except Exception as e:
                errors.append(f"Cannot create output directory: {e}")
        
        # Check filename
        if not config.filename:
            errors.append("Filename is required")
        
        return errors
    
    def _create_output_path(self, config: ExportConfig) -> Path:
        """Create output file path.
        
        Args:
            config: Export configuration
            
        Returns:
            Path: Output file path
        """
        filename = config.filename
        if not filename.endswith(f".{self.format.value}"):
            filename += f".{self.format.value}"
        
        return config.output_directory / filename


class TextExporter(BaseExporter):
    """Text format exporter."""
    
    def __init__(self):
        """Initialize text exporter."""
        super().__init__(ExportFormat.TXT)
        self.supported_data_types = [DataType.TEXT, DataType.LIST]
    
    def export(self, data: ExportData, config: ExportConfig) -> ExportResult:
        """Export to text format.
        
        Args:
            data: Data to export
            config: Export configuration
            
        Returns:
            ExportResult: Export result
        """
        result = ExportResult(export_config=config, export_data=data)
        result.status = ExportStatus.PROCESSING
        
        try:
            output_path = self._create_output_path(config)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                # Write title
                if data.title:
                    f.write(f"{data.title}\n")
                    f.write("=" * len(data.title) + "\n\n")
                
                # Write metadata
                if config.include_metadata:
                    if data.author:
                        f.write(f"Author: {data.author}\n")
                    if config.include_timestamps:
                        f.write(f"Created: {data.created_date.strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write("\n")
                
                # Write content
                if data.data_type == DataType.TEXT:
                    f.write(str(data.content))
                elif data.data_type == DataType.LIST:
                    for i, item in enumerate(data.content, 1):
                        f.write(f"{i}. {item}\n")
                
                # Write sections
                for section in data.sections:
                    f.write(f"\n\n{section.get('title', 'Section')}\n")
                    f.write("-" * len(section.get('title', 'Section')) + "\n")
                    f.write(f"{section.get('content', '')}\n")
            
            result.output_path = output_path
            result.file_size = output_path.stat().st_size
            result.status = ExportStatus.COMPLETED
            result.success = True
            result.items_exported = 1
            
        except Exception as e:
            result.status = ExportStatus.FAILED
            result.error_message = str(e)
        
        result.end_time = datetime.now()
        result.duration = (result.end_time - result.start_time).total_seconds()
        
        return result


class CSVExporter(BaseExporter):
    """CSV format exporter."""
    
    def __init__(self):
        """Initialize CSV exporter."""
        super().__init__(ExportFormat.CSV)
        self.supported_data_types = [DataType.TABLE, DataType.LIST]
    
    def export(self, data: ExportData, config: ExportConfig) -> ExportResult:
        """Export to CSV format.
        
        Args:
            data: Data to export
            config: Export configuration
            
        Returns:
            ExportResult: Export result
        """
        result = ExportResult(export_config=config, export_data=data)
        result.status = ExportStatus.PROCESSING
        
        try:
            output_path = self._create_output_path(config)
            
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                # Write headers if enabled
                if config.include_headers and data.data_type == DataType.TABLE:
                    if isinstance(data.content, dict) and 'headers' in data.content:
                        writer.writerow(data.content['headers'])
                
                # Write data
                if data.data_type == DataType.TABLE:
                    if isinstance(data.content, dict) and 'rows' in data.content:
                        for row in data.content['rows']:
                            writer.writerow(row)
                    elif isinstance(data.content, list):
                        for row in data.content:
                            if isinstance(row, (list, tuple)):
                                writer.writerow(row)
                            else:
                                writer.writerow([row])
                elif data.data_type == DataType.LIST:
                    for item in data.content:
                        writer.writerow([item])
            
            result.output_path = output_path
            result.file_size = output_path.stat().st_size
            result.status = ExportStatus.COMPLETED
            result.success = True
            result.items_exported = len(data.content) if isinstance(data.content, list) else 1
            
        except Exception as e:
            result.status = ExportStatus.FAILED
            result.error_message = str(e)
        
        result.end_time = datetime.now()
        result.duration = (result.end_time - result.start_time).total_seconds()
        
        return result


class JSONExporter(BaseExporter):
    """JSON format exporter."""
    
    def __init__(self):
        """Initialize JSON exporter."""
        super().__init__(ExportFormat.JSON)
        self.supported_data_types = [DataType.TEXT, DataType.TABLE, DataType.LIST, DataType.MIXED]
    
    def export(self, data: ExportData, config: ExportConfig) -> ExportResult:
        """Export to JSON format.
        
        Args:
            data: Data to export
            config: Export configuration
            
        Returns:
            ExportResult: Export result
        """
        result = ExportResult(export_config=config, export_data=data)
        result.status = ExportStatus.PROCESSING
        
        try:
            output_path = self._create_output_path(config)
            
            # Prepare data structure
            export_dict = {
                "title": data.title,
                "content": data.content,
                "data_type": data.data_type.value
            }
            
            # Add metadata if enabled
            if config.include_metadata:
                export_dict["metadata"] = {
                    "author": data.author,
                    "subject": data.subject,
                    "keywords": data.keywords,
                    "created_date": data.created_date.isoformat() if config.include_timestamps else None
                }
            
            # Add sections
            if data.sections:
                export_dict["sections"] = data.sections
            
            # Add custom fields
            if data.custom_fields:
                export_dict["custom_fields"] = data.custom_fields
            
            # Write JSON file
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(export_dict, f, indent=2, ensure_ascii=False, default=str)
            
            result.output_path = output_path
            result.file_size = output_path.stat().st_size
            result.status = ExportStatus.COMPLETED
            result.success = True
            result.items_exported = 1
            
        except Exception as e:
            result.status = ExportStatus.FAILED
            result.error_message = str(e)
        
        result.end_time = datetime.now()
        result.duration = (result.end_time - result.start_time).total_seconds()
        
        return result


class PDFExporter(BaseExporter):
    """PDF format exporter."""
    
    def __init__(self):
        """Initialize PDF exporter."""
        super().__init__(ExportFormat.PDF)
        self.supported_data_types = [DataType.TEXT, DataType.TABLE, DataType.LIST, DataType.MIXED]
    
    def export(self, data: ExportData, config: ExportConfig) -> ExportResult:
        """Export to PDF format.
        
        Args:
            data: Data to export
            config: Export configuration
            
        Returns:
            ExportResult: Export result
        """
        result = ExportResult(export_config=config, export_data=data)
        result.status = ExportStatus.PROCESSING
        
        if SimpleDocTemplate is None:
            result.status = ExportStatus.FAILED
            result.error_message = "ReportLab not available for PDF export"
            return result
        
        try:
            output_path = self._create_output_path(config)
            
            # Create PDF document
            doc = SimpleDocTemplate(
                str(output_path),
                pagesize=letter if config.page_size == "Letter" else A4
            )
            
            # Build content
            story = []
            styles = getSampleStyleSheet()
            
            # Title
            if data.title:
                title_para = Paragraph(data.title, styles['Title'])
                story.append(title_para)
                story.append(Spacer(1, 12))
            
            # Metadata
            if config.include_metadata:
                if data.author:
                    author_para = Paragraph(f"<b>Author:</b> {data.author}", styles['Normal'])
                    story.append(author_para)
                
                if config.include_timestamps:
                    date_para = Paragraph(
                        f"<b>Created:</b> {data.created_date.strftime('%Y-%m-%d %H:%M:%S')}",
                        styles['Normal']
                    )
                    story.append(date_para)
                
                story.append(Spacer(1, 12))
            
            # Content
            if data.data_type == DataType.TEXT:
                content_para = Paragraph(str(data.content), styles['Normal'])
                story.append(content_para)
            
            elif data.data_type == DataType.LIST:
                for item in data.content:
                    item_para = Paragraph(f"• {item}", styles['Normal'])
                    story.append(item_para)
            
            elif data.data_type == DataType.TABLE:
                if isinstance(data.content, dict) and 'rows' in data.content:
                    table_data = data.content['rows']
                    if config.include_headers and 'headers' in data.content:
                        table_data = [data.content['headers']] + table_data
                    
                    table = Table(table_data)
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 14),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    story.append(table)
            
            # Sections
            for section in data.sections:
                story.append(Spacer(1, 12))
                section_title = Paragraph(section.get('title', 'Section'), styles['Heading2'])
                story.append(section_title)
                
                section_content = Paragraph(section.get('content', ''), styles['Normal'])
                story.append(section_content)
            
            # Build PDF
            doc.build(story)
            
            result.output_path = output_path
            result.file_size = output_path.stat().st_size
            result.status = ExportStatus.COMPLETED
            result.success = True
            result.pages_exported = 1  # Would need to calculate actual pages
            result.items_exported = 1
            
        except Exception as e:
            result.status = ExportStatus.FAILED
            result.error_message = str(e)
        
        result.end_time = datetime.now()
        result.duration = (result.end_time - result.start_time).total_seconds()
        
        return result


class ExcelExporter(BaseExporter):
    """Excel format exporter."""
    
    def __init__(self):
        """Initialize Excel exporter."""
        super().__init__(ExportFormat.XLSX)
        self.supported_data_types = [DataType.TABLE, DataType.LIST, DataType.MIXED]
    
    def export(self, data: ExportData, config: ExportConfig) -> ExportResult:
        """Export to Excel format.
        
        Args:
            data: Data to export
            config: Export configuration
            
        Returns:
            ExportResult: Export result
        """
        result = ExportResult(export_config=config, export_data=data)
        result.status = ExportStatus.PROCESSING
        
        if Workbook is None:
            result.status = ExportStatus.FAILED
            result.error_message = "openpyxl not available for Excel export"
            return result
        
        try:
            output_path = self._create_output_path(config)
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = data.title or "Data"
            
            row = 1
            
            # Title
            if data.title:
                ws.cell(row=row, column=1, value=data.title)
                ws.cell(row=row, column=1).font = Font(size=16, bold=True)
                row += 2
            
            # Metadata
            if config.include_metadata:
                if data.author:
                    ws.cell(row=row, column=1, value="Author:")
                    ws.cell(row=row, column=2, value=data.author)
                    row += 1
                
                if config.include_timestamps:
                    ws.cell(row=row, column=1, value="Created:")
                    ws.cell(row=row, column=2, value=data.created_date.strftime('%Y-%m-%d %H:%M:%S'))
                    row += 1
                
                row += 1
            
            # Content
            if data.data_type == DataType.TABLE:
                if isinstance(data.content, dict):
                    # Headers
                    if config.include_headers and 'headers' in data.content:
                        for col, header in enumerate(data.content['headers'], 1):
                            cell = ws.cell(row=row, column=col, value=header)
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                        row += 1
                    
                    # Rows
                    if 'rows' in data.content:
                        for data_row in data.content['rows']:
                            for col, value in enumerate(data_row, 1):
                                ws.cell(row=row, column=col, value=value)
                            row += 1
            
            elif data.data_type == DataType.LIST:
                for item in data.content:
                    ws.cell(row=row, column=1, value=item)
                    row += 1
            
            # Sections as separate sheets
            for section in data.sections:
                section_ws = wb.create_sheet(title=section.get('title', 'Section'))
                section_ws.cell(row=1, column=1, value=section.get('content', ''))
            
            # Save workbook
            wb.save(output_path)
            
            result.output_path = output_path
            result.file_size = output_path.stat().st_size
            result.status = ExportStatus.COMPLETED
            result.success = True
            result.items_exported = len(data.content) if isinstance(data.content, list) else 1
            
        except Exception as e:
            result.status = ExportStatus.FAILED
            result.error_message = str(e)
        
        result.end_time = datetime.now()
        result.duration = (result.end_time - result.start_time).total_seconds()
        
        return result


class ExportSystem:
    """Main export system manager."""
    
    def __init__(self):
        """Initialize export system."""
        # Exporters registry
        self.exporters: Dict[ExportFormat, BaseExporter] = {
            ExportFormat.TXT: TextExporter(),
            ExportFormat.CSV: CSVExporter(),
            ExportFormat.JSON: JSONExporter(),
            ExportFormat.PDF: PDFExporter(),
            ExportFormat.XLSX: ExcelExporter()
        }
        
        # Export history
        self.export_history: List[ExportResult] = []
        
        # Event handlers
        self.event_handlers: Dict[str, List[Callable]] = {
            "export_started": [],
            "export_completed": [],
            "export_failed": [],
            "export_progress": []
        }
    
    def register_exporter(self, format: ExportFormat, exporter: BaseExporter):
        """Register custom exporter.
        
        Args:
            format: Export format
            exporter: Exporter instance
        """
        self.exporters[format] = exporter
    
    def get_supported_formats(self) -> List[ExportFormat]:
        """Get supported export formats.
        
        Returns:
            List[ExportFormat]: Supported formats
        """
        return list(self.exporters.keys())
    
    def can_export(self, format: ExportFormat, data_type: DataType) -> bool:
        """Check if format can export data type.
        
        Args:
            format: Export format
            data_type: Data type
            
        Returns:
            bool: True if supported
        """
        if format in self.exporters:
            return self.exporters[format].can_export(data_type)
        return False
    
    def export_data(self, data: ExportData, config: ExportConfig) -> ExportResult:
        """Export data to specified format.
        
        Args:
            data: Data to export
            config: Export configuration
            
        Returns:
            ExportResult: Export result
        """
        # Validate format
        if config.format not in self.exporters:
            result = ExportResult(export_config=config, export_data=data)
            result.status = ExportStatus.FAILED
            result.error_message = f"Unsupported export format: {config.format}"
            return result
        
        exporter = self.exporters[config.format]
        
        # Validate data type
        if not exporter.can_export(data.data_type):
            result = ExportResult(export_config=config, export_data=data)
            result.status = ExportStatus.FAILED
            result.error_message = f"Format {config.format} cannot export {data.data_type} data"
            return result
        
        # Validate configuration
        validation_errors = exporter.validate_config(config)
        if validation_errors:
            result = ExportResult(export_config=config, export_data=data)
            result.status = ExportStatus.FAILED
            result.error_message = "; ".join(validation_errors)
            return result
        
        # Emit start event
        self._emit_event("export_started", {"config": config, "data": data})
        
        # Perform export
        result = exporter.export(data, config)
        
        # Add to history
        self.export_history.append(result)
        
        # Emit completion event
        if result.success:
            self._emit_event("export_completed", {"result": result})
        else:
            self._emit_event("export_failed", {"result": result})
        
        return result
    
    def export_multiple(self, exports: List[Tuple[ExportData, ExportConfig]]) -> List[ExportResult]:
        """Export multiple data sets.
        
        Args:
            exports: List of (data, config) tuples
            
        Returns:
            List[ExportResult]: Export results
        """
        results = []
        
        for i, (data, config) in enumerate(exports):
            # Emit progress
            self._emit_event("export_progress", {
                "current": i + 1,
                "total": len(exports),
                "percentage": (i + 1) / len(exports) * 100
            })
            
            result = self.export_data(data, config)
            results.append(result)
        
        return results
    
    def create_archive(self, files: List[Path], archive_path: Path, format: str = "zip") -> bool:
        """Create archive from multiple files.
        
        Args:
            files: Files to archive
            archive_path: Output archive path
            format: Archive format (zip, tar)
            
        Returns:
            bool: True if successful
        """
        try:
            if format.lower() == "zip":
                with zipfile.ZipFile(archive_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    for file_path in files:
                        if file_path.exists():
                            zipf.write(file_path, file_path.name)
                return True
            else:
                # Could implement tar support here
                return False
                
        except Exception as e:
            print(f"Error creating archive: {e}")
            return False
    
    def get_export_history(self, limit: int = None) -> List[ExportResult]:
        """Get export history.
        
        Args:
            limit: Maximum number of results
            
        Returns:
            List[ExportResult]: Export history
        """
        history = sorted(self.export_history, key=lambda x: x.start_time, reverse=True)
        
        if limit:
            history = history[:limit]
        
        return history
    
    def clear_history(self):
        """Clear export history."""
        self.export_history.clear()
    
    def _emit_event(self, event: str, data: Any):
        """Emit event to handlers.
        
        Args:
            event: Event name
            data: Event data
        """
        if event in self.event_handlers:
            for handler in self.event_handlers[event]:
                try:
                    handler(event, data)
                except Exception as e:
                    print(f"Error in event handler: {e}")
    
    def add_event_handler(self, event: str, handler: Callable):
        """Add event handler.
        
        Args:
            event: Event name
            handler: Event handler function
        """
        if event in self.event_handlers:
            if handler not in self.event_handlers[event]:
                self.event_handlers[event].append(handler)
    
    def remove_event_handler(self, event: str, handler: Callable):
        """Remove event handler.
        
        Args:
            event: Event name
            handler: Event handler function
        """
        if event in self.event_handlers:
            if handler in self.event_handlers[event]:
                self.event_handlers[event].remove(handler)


# Global export system instance
_export_system: Optional[ExportSystem] = None


def get_export_system() -> Optional[ExportSystem]:
    """Get global export system instance.
    
    Returns:
        Optional[ExportSystem]: Global export system or None
    """
    return _export_system


def set_export_system(system: ExportSystem):
    """Set global export system instance.
    
    Args:
        system: Export system to set
    """
    global _export_system
    _export_system = system


# Convenience functions
def initialize_export_system() -> ExportSystem:
    """Initialize export system (convenience function).
    
    Returns:
        ExportSystem: Initialized export system
    """
    global _export_system
    _export_system = ExportSystem()
    return _export_system


def export_to_file(data: Any, file_path: Path, format: ExportFormat = None, **kwargs) -> bool:
    """Export data to file (convenience function).
    
    Args:
        data: Data to export
        file_path: Output file path
        format: Export format (auto-detected if None)
        **kwargs: Additional configuration options
        
    Returns:
        bool: True if successful
    """
    if not _export_system:
        initialize_export_system()
    
    # Auto-detect format from file extension
    if format is None:
        extension = file_path.suffix[1:].lower()
        try:
            format = ExportFormat(extension)
        except ValueError:
            print(f"Unsupported file extension: {extension}")
            return False
    
    # Create export data
    export_data = ExportData(
        title=kwargs.get('title', file_path.stem),
        content=data,
        data_type=kwargs.get('data_type', DataType.TEXT)
    )
    
    # Create export config
    config = ExportConfig(
        format=format,
        filename=file_path.stem,
        output_directory=file_path.parent
    )
    
    # Apply additional config options
    for key, value in kwargs.items():
        if hasattr(config, key):
            setattr(config, key, value)
    
    # Export
    result = _export_system.export_data(export_data, config)
    return result.success


def export_to_pdf(data: Any, file_path: Path, title: str = "", **kwargs) -> bool:
    """Export data to PDF (convenience function).
    
    Args:
        data: Data to export
        file_path: Output PDF path
        title: Document title
        **kwargs: Additional options
        
    Returns:
        bool: True if successful
    """
    return export_to_file(
        data=data,
        file_path=file_path,
        format=ExportFormat.PDF,
        title=title,
        **kwargs
    )


def export_to_excel(data: Any, file_path: Path, title: str = "", **kwargs) -> bool:
    """Export data to Excel (convenience function).
    
    Args:
        data: Data to export
        file_path: Output Excel path
        title: Worksheet title
        **kwargs: Additional options
        
    Returns:
        bool: True if successful
    """
    return export_to_file(
        data=data,
        file_path=file_path,
        format=ExportFormat.XLSX,
        title=title,
        data_type=DataType.TABLE,
        **kwargs
    )